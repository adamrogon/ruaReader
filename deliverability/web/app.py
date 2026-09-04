"""Dashboard application.

Deliberately a small server-rendered app rather than an SPA: templates for the
pages, a handful of JSON endpoints for the charts.

The overview is built to answer one question quickly — "does any domain have a
problem today, and with which provider" — so it opens on the urgency-ordered
domain list with the per-ESP breakdown one click away.
"""

from __future__ import annotations

import datetime as dt
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import JSONResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from .. import jobs
from ..classify.esp import ESP_DISPLAY_ORDER
from ..classify.esp import OTHER as ESP_OTHER
from ..classify.esp import UNKNOWN as ESP_UNKNOWN
from ..classify.esp import esp_from_org_name
from ..config import Mailbox, Settings, load_domains
from ..health import STALENESS_HOURS, domain_statuses, ingestion_health, localize_ingestion_health
from ..insights import summarize_domain
from ..i18n import DEFAULT_LANG, SUPPORTED_LANGS, resolve_lang, translate
from ..secrets import SecretsError
from ..secrets import encrypt as encrypt_secret
from ..secrets import is_configured as secret_key_configured
from ..storage import (
    BlacklistRepository,
    BounceRepository,
    DismissedFlagRepository,
    DmarcRepository,
    DnsRepository,
    DomainConfigRepository,
    FolderRepository,
    MailboxConfigRepository,
    get_database,
)
from ..validate import test_domain, test_mailbox

BASE_DIR = Path(__file__).resolve().parent

app = FastAPI(title="Deliverability Monitor", docs_url=None, redoc_url=None)
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def _bold_markers(text: str) -> "Any":
    """Turn markdown-ish markers in translated strings into inline HTML,
    and split "description" from "what to do" into two paragraphs.

    Every flag message is one description paragraph, optionally followed by a
    ``\\n\\n``-separated "what to do" paragraph. Rendering both as one blob of
    ``pre-line`` text worked, but the two roles blurred into each other; a
    second ``<p>`` with its own class (with the callout tint added in CSS)
    turns them into visually separate blocks — same content, clearer purpose.

    Runs AFTER Jinja/markupsafe has escaped the rendered string, so any HTML
    that leaked into a parameter (a mailbox name someone typed, a diagnostic
    string from a mail server) has already been neutralised — only the literal
    ``**`` and `` ` `` markers, which the escaper leaves alone, are turned
    into tags. That is what makes it safe to mark the result with
    :func:`Markup` here.
    """
    import re

    from markupsafe import Markup, escape

    result = str(escape(text))
    result = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", result)
    result = re.sub(r"`([^`]+?)`", r"<code>\1</code>", result)

    # Split on the first blank line: everything before is the description,
    # everything after is the action. maxsplit=1 keeps subsequent \n\n as
    # literal breaks in the action paragraph (rare, but preserved).
    parts = re.split(r"\n{2,}", result, maxsplit=1)
    if len(parts) == 2:
        html = f'<p class="flag-desc">{parts[0].strip()}</p>' \
               f'<p class="flag-action">{parts[1].strip()}</p>'
    else:
        html = f'<p class="flag-desc">{parts[0].strip()}</p>'
    return Markup(html)


templates.env.filters["bold_markers"] = _bold_markers


def _markdown_to_html(text: str) -> str:
    """Turn **bold** and blank-line-separated paragraphs into HTML.

    Used for the AI summary — unlike ``_bold_markers`` this keeps every
    paragraph (an AI summary is a handful of separate points, not one fixed
    description+action pair), but the same safety property holds: the text
    is HTML-escaped first, so only the literal ``**`` markers the escaper
    leaves alone become tags. The model output is plain sentences built by
    ``insights.py`` from pre-verified facts — this is purely a readability
    pass, not a trust boundary.
    """
    import re

    from markupsafe import escape

    result = str(escape(text))
    # Belt-and-suspenders: the prompt asks the model not to add its own
    # title, but strip a stray leading "# Heading" line if one shows up
    # anyway rather than rendering a literal "#" in the card.
    result = re.sub(r"^#{1,6}\s*(.+)$", r"\1", result, flags=re.MULTILINE)
    result = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", result)
    paragraphs = [p.strip() for p in re.split(r"\n{2,}", result) if p.strip()]
    return "".join(f"<p>{p}</p>" for p in paragraphs) or f"<p>{result}</p>"


def _static_version() -> str:
    """Modtime-based cache-buster for the stylesheet URL.

    Browsers were caching the CSS aggressively across every restart. Appending
    ``?v=<mtime>`` to the ``<link>`` href forces a fresh fetch whenever the
    file actually changes, and — importantly — costs the browser nothing when
    the file has NOT changed (the URL is identical, so its cache hit still
    works). Computed on import, which is fine for a local single-user app.
    """
    try:
        return str(int((BASE_DIR / "static" / "style.css").stat().st_mtime))
    except OSError:
        return "0"


templates.env.globals["static_version"] = _static_version()


# Canonical ESP name (see classify/esp.py) -> logo file in static/logos/.
# These are each provider's own real favicon (fetched once from
# icons.duckduckgo.com and normalised to 64x64 PNG), not a redrawn icon set —
# multi-color and immediately recognisable, unlike a single-tint glyph.
# Anything not listed here (Comcast, Other, Unknown) falls back to the plain
# colored-initial swatch in the template.
_ESP_LOGO_FILES = {
    "Google": "google.png",
    "Microsoft": "microsoft.png",
    "Yahoo": "yahoo.png",
    "WP/O2": "o2.png",
    "Mail.ru": "mailru.png",
    "Proton": "protonmail.png",
    "Zoho": "zoho.png",
    "Apple": "apple.png",
    "AOL": "aol.png",
    "GMX/United Internet": "gmx.png",
    "Fastmail": "fastmail.png",
    "Seznam": "seznam.png",
    "Onet": "onet.png",
    "Interia": "interia.png",
}


def _esp_logo(esp_name: Optional[str]) -> str:
    """Static URL for an ESP's logo, or '' if none is on file."""
    filename = _ESP_LOGO_FILES.get(esp_name or "")
    return f"/static/logos/{filename}" if filename else ""


templates.env.globals["esp_logo"] = _esp_logo


def _bounce_code_help(code: Optional[str], lang: str) -> str:
    """Human one-liner for an SMTP status code, or '' if none is defined.

    Returned as an empty string (not the raw key) when no translation exists,
    so the template can just check truthiness before rendering the help row.
    """
    if not code:
        return ""
    key = f"bounce.code_help.{code}"
    resolved = translate(key, lang)
    return "" if resolved == key else resolved


def _ip_hint(ptr_hostname: Optional[str]) -> Optional[str]:
    """A short, honest identification hint for a blacklist-checked IP.

    Built only from the PTR hostname already captured at ingest time (see
    ``ingest/blacklist.py``) — never a live lookup here, since this runs on
    every page render. Shows the raw hostname, and when it matches a known
    provider's pattern (the same table used to label ESPs elsewhere in the
    app), appends that provider's friendly name too — e.g.
    "mail123.ovh.net (OVH)". Returns None when there's no PTR record to show.
    """
    if not ptr_hostname:
        return None
    label = esp_from_org_name(ptr_hostname)
    if label in (ESP_OTHER, ESP_UNKNOWN):
        return ptr_hostname
    return f"{ptr_hostname} ({label})"


def _trim_diagnostic(text: Optional[str], max_len: int = 250) -> str:
    """Extract the meaningful part of a bounce diagnostic.

    Real bounces routinely include the original message body — signature,
    disclaimers, the outreach pitch itself. This function finds the actual
    bounce content by anchoring on strong markers (``failed:``, an enhanced
    status code like ``5.7.1``, ``smtp;``, ``rejected``) and keeps a window
    of text around them. Falls back to a plain truncation when no anchor
    exists so a row is never returned empty.

    Deliberately excludes weaker words like ``delivery`` (matches "delivery
    software" too), ``mailbox``, ``recipient`` — those cause false positives
    in the ``This message was created automatically by mail delivery
    software`` preamble that most bounces start with.
    """
    import re

    if not text:
        return ""
    text = text.strip()

    # Where the useful bounce content typically begins.
    anchor = re.compile(
        r"\b[245]\.\d+\.\d+\b|failed:|smtp\s*;|\brejected\b|address(?:es)?\s+failed",
        re.IGNORECASE,
    )
    # Where the useful content typically ENDS — a greeting, a signature
    # separator, a quoted-text prefix, a horizontal divider. Everything past
    # one of these is the outreach body echoed back into the bounce, not
    # something the reader needs to see in the summary.
    end_marker = re.compile(
        r"\n\s*(?:Hello|Hi|Dear|Cześć|Witam|Dzień\s+dobry)[, ]|"
        r"\n--\s*\n|\n\s*[_=─-]{6,}|\n\s*>|"
        r"\n\s*(?:Best\s+regards|Kind\s+regards|Regards|Sincerely|Thanks|Pozdrawiam|Pozdrowienia)\b",
        re.IGNORECASE,
    )

    m = anchor.search(text)
    if m:
        start = max(0, m.start() - 40)
        # Prefer cutting at an end-marker if there is one within the window.
        end_search = end_marker.search(text, m.end())
        hard_end = start + max_len
        if end_search and end_search.start() < hard_end:
            end = end_search.start()
        else:
            end = hard_end
        return ("…" if start else "") + text[start:end].rstrip() + ("…" if end < len(text) else "")

    # No anchor: return the text as-is if short, otherwise a plain truncation.
    if len(text) <= max_len:
        return text
    return text[:max_len].rstrip() + "…"

DEFAULT_WINDOW_DAYS = 7
LANG_COOKIE = "lang"


def _context() -> Dict[str, Any]:
    settings = Settings.from_env()
    return {"settings": settings, "database": get_database(settings)}


def _resolve_request_lang(request: Request) -> str:
    """?lang= wins, falling back to the cookie set by a previous visit."""
    requested = request.query_params.get("lang") or request.cookies.get(LANG_COOKIE)
    return resolve_lang(requested)


def _render(request: Request, template: str, lang: str, extra: Dict[str, Any]) -> Response:
    """TemplateResponse plus the i18n context every page needs, and a cookie
    so the chosen language survives a visit that doesn't carry ?lang=."""
    other_lang = next(candidate for candidate in SUPPORTED_LANGS if candidate != lang)
    context = {
        "lang": lang,
        "other_lang": other_lang,
        "t": lambda key, **kw: translate(key, lang, **kw),
        **extra,
    }
    response = templates.TemplateResponse(request, template, context)
    response.set_cookie(LANG_COOKIE, lang, max_age=60 * 60 * 24 * 365, samesite="lax")
    return response


def _since(days: int) -> dt.datetime:
    return dt.datetime.now(dt.timezone.utc) - dt.timedelta(days=days)


def _day_key(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)[:10]


def _date_axis(days: int) -> List[str]:
    """A continuous date axis, so gaps in reporting are visible as gaps."""
    today = dt.datetime.now(dt.timezone.utc).date()
    return [(today - dt.timedelta(days=offset)).isoformat() for offset in range(days - 1, -1, -1)]


def _overview_context(
    request: Request, days: int, lang: str, domains: Optional[List[Any]] = None
) -> Dict[str, Any]:
    """Shared by "/" (every domain) and "/folder/{id}" (one folder's domains)
    — the only difference is which domains get passed to domain_statuses().
    """
    ctx = _context()
    raw_statuses = domain_statuses(ctx["settings"], ctx["database"], domains=domains, window_days=days)
    # Resolved to the active language once here; nothing downstream needs to
    # know a second language exists.
    statuses = [s.localize(lang) for s in raw_statuses]
    health = localize_ingestion_health(ingestion_health(ctx["database"], ctx["settings"]), lang)

    passed_total = sum(s["metrics"].get("passed", 0) or 0 for s in statuses)
    evaluated_total = sum(
        (s["metrics"].get("passed", 0) or 0) + (s["metrics"].get("failed", 0) or 0) for s in statuses
    )
    totals = {
        "domains": len(statuses),
        "critical": sum(1 for s in statuses if s["severity"] == "critical"),
        "warning": sum(1 for s in statuses if s["severity"] == "warning"),
        "ok": sum(1 for s in statuses if s["severity"] in ("ok", "info")),
        "sender_blocks": sum(s["metrics"].get("bounces_sender_block", 0) for s in statuses),
        "blacklisted": sum(1 for s in statuses if s["metrics"].get("critical_blacklist")),
        "messages": sum(s["metrics"].get("messages", 0) for s in statuses),
        "compliance": (passed_total / evaluated_total) if evaluated_total else None,
        "bounces_total": sum(s["metrics"].get("bounces_total", 0) for s in statuses),
        "spf_near_limit": sum(
            1
            for s in statuses
            if s["metrics"].get("spf_lookups") is not None and s["metrics"]["spf_lookups"] >= 8
        ),
    }

    # Fleet-wide per-ESP totals — the "is it Google or Yahoo" question asked
    # across every domain at once.
    esp_totals: Dict[str, Dict[str, Any]] = defaultdict(
        lambda: {"esp": "", "pass": 0, "failed": 0, "forwarded": 0, "domains": set()}
    )
    for status in statuses:
        for row in status["esp_rows"]:
            entry = esp_totals[row["esp"]]
            entry["esp"] = row["esp"]
            entry["pass"] += row["pass"]
            entry["failed"] += row["failed"]
            entry["forwarded"] += row["forwarded"]
            if row["failed"]:
                entry["domains"].add(status["domain"])

    esp_summary = []
    for entry in esp_totals.values():
        evaluated = entry["pass"] + entry["failed"]
        esp_summary.append(
            {
                "esp": entry["esp"],
                "pass": entry["pass"],
                "failed": entry["failed"],
                "forwarded": entry["forwarded"],
                "total": evaluated + entry["forwarded"],
                "evaluated": evaluated,
                "compliance": (entry["pass"] / evaluated) if evaluated else None,
                "affected_domains": sorted(entry["domains"]),
            }
        )
    esp_summary.sort(key=lambda e: (ESP_DISPLAY_ORDER.get(e["esp"], 50), -e["total"]))

    domain_filter = [d.name for d in domains] if domains is not None else None
    other_providers = DmarcRepository(ctx["database"], ctx["settings"].project_id).other_esp_org_breakdown(
        _since(days), domain_filter
    )

    return {
        "statuses": statuses,
        "health": health,
        "totals": totals,
        "esp_summary": esp_summary,
        "other_providers": other_providers,
        "days": days,
        "staleness_hours": STALENESS_HOURS,
    }


@app.get("/")
def overview(request: Request, days: int = Query(DEFAULT_WINDOW_DAYS, ge=1, le=90)):
    """Prototype: home is folder tiles, not a flat list of every domain — a
    department opens their own folder instead of scrolling past everyone
    else's. "Wszystkie domeny" is one tile among these, not the default.
    """
    lang = _resolve_request_lang(request)
    ctx = _context()
    folder_repo = FolderRepository(ctx["database"], ctx["settings"].project_id)
    domain_repo = DomainConfigRepository(ctx["database"], ctx["settings"].project_id)

    # Fleet-wide totals + trend charts, shared with "/domains" and
    # "/folder/{id}" — the home page shows the same numbers, just rolled up
    # across every folder instead of one.
    fleet = _overview_context(request, days, lang)
    status_by_domain = {s["domain"]: s for s in fleet["statuses"]}

    all_folders = folder_repo.list_all()
    counts = folder_repo.domain_counts()
    all_domains_by_id = {d["id"]: d for d in domain_repo.list_all()}

    rows = []
    for folder in all_folders:
        domain_ids = folder_repo.domain_ids_in_folder(folder["id"])
        names = {all_domains_by_id[did]["name"] for did in domain_ids if did in all_domains_by_id}
        folder_statuses = [status_by_domain[n] for n in names if n in status_by_domain]
        messages = sum(s["metrics"].get("messages", 0) for s in folder_statuses)
        passed = sum(s["metrics"].get("passed", 0) or 0 for s in folder_statuses)
        evaluated_total = sum(
            (s["metrics"].get("passed", 0) or 0) + (s["metrics"].get("failed", 0) or 0) for s in folder_statuses
        )
        # SPF lookups aren't additive across domains (each is its own count
        # against the same fixed limit of 10) — the meaningful fleet-level
        # rollup is how many domains are close to or over that limit, same
        # shape as the critical/warning domain counts below.
        spf_near_limit = sum(
            1
            for s in folder_statuses
            if s["metrics"].get("spf_lookups") is not None
            and s["metrics"]["spf_lookups"] >= 8
        )
        rows.append(
            {
                "id": folder["id"],
                "name": folder["name"],
                "domain_count": counts.get(folder["id"], 0),
                "messages": messages,
                "compliance": (passed / evaluated_total) if evaluated_total else None,
                "bounces_total": sum(s["metrics"].get("bounces_total", 0) for s in folder_statuses),
                "spf_near_limit": spf_near_limit,
                "critical": sum(1 for s in folder_statuses if s["severity"] == "critical"),
                "warning": sum(1 for s in folder_statuses if s["severity"] == "warning"),
                "ok": sum(1 for s in folder_statuses if s["severity"] in ("ok", "info")),
            }
        )

    context = dict(fleet)
    context["rows"] = rows
    return _render(request, "folders.html", lang, context)


@app.get("/domains")
def all_domains(request: Request, days: int = Query(DEFAULT_WINDOW_DAYS, ge=1, le=90)):
    """The old "/" — every domain, flat, no folder grouping. Reached from the
    "Wszystkie domeny" tile on the new home page, not from the sidebar.
    """
    lang = _resolve_request_lang(request)
    return _render(request, "overview.html", lang, _overview_context(request, days, lang))


@app.get("/folder/{folder_id}")
def folder_detail(request: Request, folder_id: int, days: int = Query(DEFAULT_WINDOW_DAYS, ge=1, le=90)):
    lang = _resolve_request_lang(request)
    ctx = _context()
    folder_repo = FolderRepository(ctx["database"], ctx["settings"].project_id)
    domain_repo = DomainConfigRepository(ctx["database"], ctx["settings"].project_id)

    folder = folder_repo.get(folder_id)
    if not folder:
        raise HTTPException(status_code=404, detail=f"Folder {folder_id} not found")

    domain_ids = set(folder_repo.domain_ids_in_folder(folder_id))
    names = {row["name"] for row in domain_repo.list_all() if row["id"] in domain_ids}
    scoped_domains = [d for d in load_domains() if d.name in names]

    context = _overview_context(request, days, lang, domains=scoped_domains)
    context["folder"] = folder
    return _render(request, "overview.html", lang, context)


@app.post("/settings/folders")
def create_folder(request: Request, name: str = Form(...)):
    lang = _resolve_request_lang(request)
    ctx = _context()
    if name.strip():
        FolderRepository(ctx["database"], ctx["settings"].project_id).create(name)
    return _settings_redirect(lang, notice="folder_saved")


@app.post("/settings/folders/{folder_id}/delete")
def delete_folder(request: Request, folder_id: int):
    lang = _resolve_request_lang(request)
    ctx = _context()
    FolderRepository(ctx["database"], ctx["settings"].project_id).delete(folder_id)
    return _settings_redirect(lang, notice="folder_deleted")


@app.post("/settings/domains/{domain_id}/folders")
def set_domain_folders(request: Request, domain_id: int, folder_ids: List[int] = Form(default=[])):
    lang = _resolve_request_lang(request)
    ctx = _context()
    FolderRepository(ctx["database"], ctx["settings"].project_id).set_domain_folders(domain_id, folder_ids)
    return _settings_redirect(lang, notice="domain_folders_saved")


@app.get("/domain/{domain}")
def domain_detail(
    request: Request,
    domain: str,
    days: int = Query(DEFAULT_WINDOW_DAYS, ge=1, le=90),
    bounce_page: int = Query(1, ge=1, le=10_000),
    # Sort/order for the raw log. Validated inside the repository against a
    # closed whitelist, so passing an unknown column here falls back to
    # newest-first rather than raising.
    bounce_sort: str = Query("received_at"),
    bounce_order: str = Query("desc"),
):
    lang = _resolve_request_lang(request)
    ctx = _context()
    settings, database = ctx["settings"], ctx["database"]

    configured = {d.name: d for d in load_domains()}
    if domain not in configured:
        raise HTTPException(status_code=404, detail=f"{domain} is not in config/domains.yml")

    raw_statuses = domain_statuses(settings, database, domains=[configured[domain]], window_days=days)
    status = raw_statuses[0].localize(lang)

    dns_row = DnsRepository(database, settings.project_id).latest_per_domain().get(domain)
    bounce_repo = BounceRepository(database, settings.project_id)
    since = _since(days)

    # Full IP list for the blacklist section — the flag message up top only
    # names the first 3 IPs (see health.py), this is the untruncated source
    # of truth for actually filing delisting requests. Listed IPs first, same
    # "most urgent first" convention as everywhere else in this app.
    blacklist_rows = [
        dict(row, hint=_ip_hint(row.get("ptr_hostname")))
        for row in sorted(
            BlacklistRepository(database, settings.project_id).latest_per_domain().get(domain, []),
            key=lambda r: (not r["listed"], r["ip"]),
        )
    ]

    # Consolidated "what's actually happening" table — folds the old separate
    # "sender blocks" and "bounce codes" panels into one row-per-signature
    # summary, each with a live sample of the diagnostic text so the SMTP
    # jargon isn't the only thing on the row.
    bounce_summary_rows = bounce_repo.summary_by_code(since, domain)
    bounce_summary = [
        dict(
            row,
            code_help=_bounce_code_help(row["status_code"], lang),
            # Full body was often megabytes of the original outreach echoed
            # back — cut to the meaningful window around the SMTP status.
            sample_diagnostic=_trim_diagnostic(row["sample_diagnostic"]),
        )
        for row in bounce_summary_rows
    ]

    # Paginated raw log — 10 per page keeps the domain page short even after
    # weeks of ingestion, and pager arithmetic runs server-side against the
    # real row count rather than being guessed from an over-fetched list.
    per_page = 10
    page_data = bounce_repo.recent_paged(
        since, domain, page=bounce_page, per_page=per_page,
        sort=bounce_sort, order=bounce_order,
    )
    recent_bounces = [
        dict(
            r,
            code_help=_bounce_code_help(r["status_code"], lang),
            diagnostic_code=_trim_diagnostic(r["diagnostic_code"]),
        )
        for r in page_data["rows"]
    ]
    total = page_data["total"]
    total_pages = max(1, (total + per_page - 1) // per_page)
    current_page = min(page_data["page"], total_pages)

    return _render(
        request,
        "domain.html",
        lang,
        {
            "status": status,
            "domain": domain,
            "dns": dns_row,
            "blacklist_rows": blacklist_rows,
            "bounce_summary": bounce_summary,
            "recent_bounces": recent_bounces,
            "bounce_pagination": {
                "current": current_page,
                "total_pages": total_pages,
                "total": total,
                "per_page": per_page,
                "has_prev": current_page > 1,
                "has_next": current_page < total_pages,
                # Echoed back the validated values so the template can build
                # sort links relative to what's actually active (not what the
                # URL happened to ask for).
                "sort": page_data["sort"],
                "order": page_data["order"],
            },
            "top_failing": DmarcRepository(database, settings.project_id).top_failing_sources(
                since, domain, limit=12
            ),
            "other_providers": DmarcRepository(database, settings.project_id).other_esp_org_breakdown(
                since, domain
            ),
            "days": days,
            "notes": configured[domain].notes,
            "selectors": configured[domain].dkim_selectors,
        },
    )


# Whitelisted so a bad ?sort= value can't do anything worse than fall back
# to the default — never build a sort expression from raw user input.
_DAY_SORT_KEYS: Dict[str, Any] = {
    "domain": lambda r: (r["domain"] or "").lower(),
    "messages": lambda r: r["messages"],
    "compliant": lambda r: r["compliant"],
    "failed": lambda r: r["failed"],
    "spf_fail": lambda r: r["spf_fail"],
    "dkim_fail": lambda r: r["dkim_fail"],
    "bounces_sender_block": lambda r: r["bounces_sender_block"],
    "bounces_hard": lambda r: r["bounces_hard"],
}


@app.get("/day/{date}")
def day_detail_fleet(
    request: Request,
    date: str,
    sort: str = Query("severity"),
    order: str = Query("desc"),
):
    """Fleet-wide drill-down for one calendar day — which domain(s) actually
    drove a spike or dip in the overview's daily chart, without opening every
    domain one at a time.
    """
    lang = _resolve_request_lang(request)
    ctx = _context()
    try:
        day = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=404, detail=f"Invalid date: {date}")

    dmarc_repo = DmarcRepository(ctx["database"], ctx["settings"].project_id)
    bounce_repo = BounceRepository(ctx["database"], ctx["settings"].project_id)

    domain_rows = dmarc_repo.domain_summary_for_day(day)
    bounce_rows = bounce_repo.counts_by_class_for_day(day)
    bounces_by_domain: Dict[str, Dict[str, int]] = defaultdict(dict)
    for row in bounce_rows:
        bounces_by_domain[row["domain"]][row["bounce_class"]] = row["count"]

    rows = []
    for row in domain_rows:
        classes = bounces_by_domain.get(row["domain"], {})
        rows.append(
            {
                **row,
                "bounces_total": sum(classes.values()),
                "bounces_sender_block": classes.get("sender_block", 0),
                "bounces_hard": classes.get("hard", 0),
            }
        )
    # Domains with only bounce activity and no DMARC records that day still
    # belong on this page.
    seen = {r["domain"] for r in rows}
    for domain_name, classes in bounces_by_domain.items():
        if domain_name not in seen:
            rows.append(
                {
                    "domain": domain_name,
                    "messages": 0,
                    "compliant": 0,
                    "failed": 0,
                    "spf_fail": 0,
                    "dkim_fail": 0,
                    "bounces_total": sum(classes.values()),
                    "bounces_sender_block": classes.get("sender_block", 0),
                    "bounces_hard": classes.get("hard", 0),
                }
            )
    if sort in _DAY_SORT_KEYS:
        rows.sort(key=_DAY_SORT_KEYS[sort], reverse=(order != "asc"))
        active_sort, active_order = sort, ("asc" if order == "asc" else "desc")
    else:
        # Default: severity — sender blocks first, then raw failures, then
        # bounce volume. Not a single column, so it isn't in the whitelist
        # above; "severity" just means "whatever ?sort= didn't override".
        rows.sort(key=lambda r: (-r["bounces_sender_block"], -r["failed"], -r["bounces_total"]))
        active_sort, active_order = "severity", "desc"

    return _render(
        request,
        "day.html",
        lang,
        {
            "day": date,
            "fleet_wide": True,
            "rows": rows,
            "sort": active_sort,
            "order": active_order,
            "back_days": DEFAULT_WINDOW_DAYS,
        },
    )


@app.get("/domain/{domain}/day/{date}")
def day_detail_domain(request: Request, domain: str, date: str):
    """Per-domain drill-down for one calendar day — the reporting orgs and
    sources behind a click on that domain's own daily chart point.
    """
    lang = _resolve_request_lang(request)
    ctx = _context()
    try:
        day = dt.date.fromisoformat(date)
    except ValueError:
        raise HTTPException(status_code=404, detail=f"Invalid date: {date}")

    dmarc_repo = DmarcRepository(ctx["database"], ctx["settings"].project_id)
    bounce_repo = BounceRepository(ctx["database"], ctx["settings"].project_id)

    records = dmarc_repo.records_for_day(domain, day)
    bounce_rows = bounce_repo.counts_by_class_for_day(day, domain=domain)
    bounce_classes = {row["bounce_class"]: row["count"] for row in bounce_rows}

    return _render(
        request,
        "day.html",
        lang,
        {
            "day": date,
            "domain": domain,
            "records": records,
            "bounce_classes": bounce_classes,
            "bounce_total": sum(bounce_classes.values()),
            "back_days": DEFAULT_WINDOW_DAYS,
        },
    )


@app.post("/domain/{domain}/flags/dismiss")
def dismiss_flag(
    request: Request,
    domain: str,
    fingerprint: str = Form(...),
    days: int = Form(DEFAULT_WINDOW_DAYS),
):
    lang = _resolve_request_lang(request)
    ctx = _context()
    DismissedFlagRepository(ctx["database"], ctx["settings"].project_id).dismiss(domain, fingerprint)
    return RedirectResponse(f"/domain/{domain}?lang={lang}&days={days}", status_code=303)


@app.post("/domain/{domain}/flags/restore")
def restore_flag(
    request: Request,
    domain: str,
    fingerprint: str = Form(...),
    days: int = Form(DEFAULT_WINDOW_DAYS),
):
    lang = _resolve_request_lang(request)
    ctx = _context()
    DismissedFlagRepository(ctx["database"], ctx["settings"].project_id).restore(domain, fingerprint)
    return RedirectResponse(f"/domain/{domain}?lang={lang}&days={days}", status_code=303)


@app.get("/api/volume")
def api_volume(domain: Optional[List[str]] = Query(None), days: int = Query(14, ge=1, le=90)) -> JSONResponse:
    """Daily DMARC message volume split by outcome."""
    ctx = _context()
    repo = DmarcRepository(ctx["database"], ctx["settings"].project_id)
    rows = repo.daily_volume(_since(days), domain)
    alignment_rows = repo.daily_alignment_failures(_since(days), domain)

    axis = _date_axis(days)
    series = {key: {day: 0 for day in axis} for key in ("pass", "forwarded", "failed")}
    for row in rows:
        day = _day_key(row["day"])
        if day in series[row["evaluation"]]:
            series[row["evaluation"]][day] += row["messages"] or 0

    # SPF Fail / DKIM Fail are alignment failures (see daily_alignment_failures'
    # docstring) — overlaid lines, not another slice of the pass/forwarded/failed
    # stack, since a message can fail alignment on one mechanism and still pass
    # DMARC overall via the other.
    spf_fail = {day: 0 for day in axis}
    dkim_fail = {day: 0 for day in axis}
    for row in alignment_rows:
        day = _day_key(row["day"])
        if day in spf_fail:
            spf_fail[day] += row["spf_fail"] or 0
            dkim_fail[day] += row["dkim_fail"] or 0

    compliance = []
    for day in axis:
        passed, failed = series["pass"][day], series["failed"][day]
        evaluated = passed + failed
        compliance.append(round(passed / evaluated * 100, 1) if evaluated else None)

    return JSONResponse(
        {
            "labels": axis,
            "passed": [series["pass"][d] for d in axis],
            "forwarded": [series["forwarded"][d] for d in axis],
            "failed": [series["failed"][d] for d in axis],
            "spf_fail": [spf_fail[d] for d in axis],
            "dkim_fail": [dkim_fail[d] for d in axis],
            "compliance": compliance,
        }
    )


@app.get("/api/bounces")
def api_bounces(domain: Optional[List[str]] = Query(None), days: int = Query(14, ge=1, le=90)) -> JSONResponse:
    """Daily bounce counts split by class."""
    ctx = _context()
    repo = BounceRepository(ctx["database"], ctx["settings"].project_id)
    rows = repo.daily_counts(_since(days), domain)

    axis = _date_axis(days)
    classes = ("sender_block", "hard", "soft", "unknown")
    series = {key: {day: 0 for day in axis} for key in classes}
    for row in rows:
        day = _day_key(row["day"])
        klass = row["bounce_class"]
        if klass in series and day in series[klass]:
            series[klass][day] += row["count"] or 0

    return JSONResponse(
        {"labels": axis, **{key: [series[key][d] for d in axis] for key in classes}}
    )


@app.get("/api/esp")
def api_esp(domain: Optional[str] = None, days: int = Query(7, ge=1, le=90)) -> JSONResponse:
    """Per-receiving-ESP outcome totals."""
    ctx = _context()
    repo = DmarcRepository(ctx["database"], ctx["settings"].project_id)
    rows = repo.esp_breakdown(_since(days), domain)

    by_esp: Dict[str, Dict[str, int]] = defaultdict(lambda: {"pass": 0, "failed": 0, "forwarded": 0})
    for row in rows:
        by_esp[row["esp"] or "Unknown"][row["evaluation"]] += row["messages"] or 0

    ordered = sorted(by_esp.items(), key=lambda kv: ESP_DISPLAY_ORDER.get(kv[0], 50))
    return JSONResponse(
        {
            "labels": [esp for esp, _ in ordered],
            "passed": [vals["pass"] for _, vals in ordered],
            "failed": [vals["failed"] for _, vals in ordered],
            "forwarded": [vals["forwarded"] for _, vals in ordered],
        }
    )


@app.post("/api/domain/{domain}/analyze")
def api_domain_analyze(
    request: Request, domain: str, days: int = Query(DEFAULT_WINDOW_DAYS, ge=1, le=90)
) -> JSONResponse:
    """"Analizuj z AI" — see deliverability/insights.py for how this stays
    grounded (the model only ever sees facts this route's own code already
    computed and verified, never raw data)."""
    lang = _resolve_request_lang(request)
    ctx = _context()
    result = summarize_domain(ctx["settings"], ctx["database"], domain, days, lang)
    result["summary_html"] = _markdown_to_html(result["summary"])
    return JSONResponse(result)


# --- Settings -----------------------------------------------------------------


def _selectors_from_form(raw: str) -> List[str]:
    """Parse the comma-separated selector field.

    An empty field means "do not check DKIM", which is different from leaving
    the domain's selectors at their defaults — so it is preserved as an empty
    list rather than silently substituted.
    """
    return [s.strip() for s in (raw or "").replace("\n", ",").split(",") if s.strip()]


def _settings_context(
    notice: Optional[str] = None, error: Optional[str] = None, bulk_results: Optional[List[Dict[str, Any]]] = None
) -> Dict[str, Any]:
    ctx = _context()
    settings, database = ctx["settings"], ctx["database"]

    domain_repo = DomainConfigRepository(database, settings.project_id)
    mailbox_repo = MailboxConfigRepository(database, settings.project_id)
    dmarc_repo = DmarcRepository(database, settings.project_id)
    folder_repo = FolderRepository(database, settings.project_id)

    mailboxes = mailbox_repo.list_all()
    domains = domain_repo.list_all()
    monitored_names = {d["name"] for d in domains}
    all_folders = folder_repo.list_all()
    folder_ids_by_domain = {d["id"]: folder_repo.folder_ids_for_domain(d["id"]) for d in domains}

    # For the bounce mailbox dropdown: the current list of choices, plus any
    # value stored on an existing bounce mailbox that is no longer on the
    # monitored list (so editing does not silently drop it). Marked as orphan
    # so the template can flag it.
    bounce_mailboxes = [m for m in mailboxes if m["kind"] == "bounce"]
    orphan_bounce_domains = sorted(
        {m["domain"] for m in bounce_mailboxes if m["domain"] and m["domain"] not in monitored_names}
    )

    return {
        "domains": domains,
        "monitored_names": sorted(monitored_names),
        "orphan_bounce_domains": orphan_bounce_domains,
        "all_folders": all_folders,
        "folder_ids_by_domain": folder_ids_by_domain,
        "rua_mailboxes": [m for m in mailboxes if m["kind"] == "rua"],
        "bounce_mailboxes": bounce_mailboxes,
        "unknown_report_domains": dmarc_repo.unknown_report_domains(),
        "secret_key_ok": secret_key_configured(),
        "notice": notice,
        "error": error,
        "days": DEFAULT_WINDOW_DAYS,
        "page": "settings",
        "bulk_results": bulk_results,
    }


@app.get("/settings")
def settings_page(request: Request, notice: Optional[str] = None, error: Optional[str] = None):
    lang = _resolve_request_lang(request)
    return _render(
        request,
        "settings.html",
        lang,
        _settings_context(notice, error),
    )


def _settings_redirect(lang: str, notice: Optional[str] = None, error: Optional[str] = None) -> RedirectResponse:
    params = [f"lang={lang}"]
    if notice:
        params.append(f"notice={notice}")
    if error:
        params.append(f"error={error}")
    # 303 so the browser re-issues as GET and a refresh does not resubmit.
    return RedirectResponse(f"/settings?{'&'.join(params)}", status_code=303)


# --- Bulk import (prototype) -------------------------------------------------
# One row = one domain + its bounce mailbox + (optionally) its rua mailbox.
# Reuses DomainConfigRepository.create() / MailboxConfigRepository.create()
# in a loop rather than any new write path — see the spec doc for why.

_BULK_IMPORT_HEADERS = [
    "domain", "dkim_selectors", "notes",
    "imap_host", "imap_port", "imap_ssl", "imap_username", "imap_password", "imap_folder",
    "same_mailbox_for_rua_and_bounce",
    "rua_imap_host", "rua_imap_port", "rua_imap_ssl", "rua_imap_username", "rua_imap_password", "rua_imap_folder",
    "folders",
]


def _truthy_yes(value: Any, default: bool = True) -> bool:
    if value is None or str(value).strip() == "":
        return default
    return str(value).strip().lower() not in ("no", "false", "0", "nie")


@app.get("/settings/bulk-import/template")
def bulk_import_template() -> Response:
    """Generated on the fly (not a static file) so the columns can never
    drift out of sync with what bulk_import() below actually parses.
    """
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Import"
    ws.append(_BULK_IMPORT_HEADERS)
    ws.append(
        [
            "example.com", "google", "Google Workspace",
            "imap.gmail.com", 993, "yes", "outreach@example.com", "haslo-aplikacji-tutaj", "INBOX",
            "yes",
            "", "", "", "", "", "",
            "Media",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions[get_column_letter(_BULK_IMPORT_HEADERS.index("same_mailbox_for_rua_and_bounce") + 1)].width = 30
    ws["J1"].comment = Comment(
        "yes = ta sama skrzynka odbiera raporty DMARC i odbicia (typowy przypadek).\n"
        "no = wypełnij też kolumny rua_imap_* osobno.",
        "Deliverability Monitor",
    )
    ws["H1"].comment = Comment("Hasło aplikacji w postaci jawnej — usuń plik z dysku po imporcie.", "Deliverability Monitor")
    for i, header in enumerate(_BULK_IMPORT_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(16, len(header) + 2)

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="szablon-import-domen.xlsx"'},
    )


@app.post("/settings/bulk-import")
async def bulk_import(request: Request, file: UploadFile = File(...)) -> Response:
    lang = _resolve_request_lang(request)
    ctx = _context()
    database, settings = ctx["database"], ctx["settings"]
    domain_repo = DomainConfigRepository(database, settings.project_id)
    mailbox_repo = MailboxConfigRepository(database, settings.project_id)
    folder_repo = FolderRepository(database, settings.project_id)

    from io import BytesIO

    from openpyxl import load_workbook

    # Read fully into memory and never touch disk — the file contains
    # plaintext IMAP passwords (see the spec doc's security section).
    raw = await file.read()
    try:
        wb = load_workbook(BytesIO(raw), data_only=True)
    except Exception:
        return _render(
            request, "settings.html", lang,
            _settings_context(error="required", bulk_results=[{"row": "-", "domain": "-", "status": "error", "reason": "Nie udało się odczytać pliku XLSX."}]),
        )
    ws = wb.active
    header_row = [str(c.value).strip() if c.value else "" for c in ws[1]]

    existing_mailbox_names = {m["name"] for m in mailbox_repo.list_all()}
    monitored_names = {d["name"] for d in domain_repo.list_all()}
    results: List[Dict[str, Any]] = []

    for row_num, raw_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row = dict(zip(header_row, raw_row))
        domain_name = str(row.get("domain") or "").strip().lower()
        if not domain_name:
            continue  # a fully blank row (e.g. trailing empty rows) is not an error, just skipped silently

        def _fail(reason: str) -> None:
            results.append({"row": row_num, "domain": domain_name, "status": "error", "reason": reason})

        try:
            if domain_name in monitored_names:
                results.append({"row": row_num, "domain": domain_name, "status": "skipped", "reason": "domena już monitorowana"})
                continue

            host = str(row.get("imap_host") or "").strip()
            username = str(row.get("imap_username") or "").strip()
            password = str(row.get("imap_password") or "").strip()
            if not host or not username or not password:
                _fail("brak imap_host / imap_username / imap_password")
                continue

            bounce_name = f"{domain_name}-bounce"
            rua_name = f"{domain_name}-rua"
            if bounce_name in existing_mailbox_names or rua_name in existing_mailbox_names:
                _fail("nazwa skrzynki już istnieje")
                continue

            same_mailbox = _truthy_yes(row.get("same_mailbox_for_rua_and_bounce"))
            if same_mailbox:
                rua_host, rua_user, rua_pass = host, username, password
                rua_port = int(row.get("imap_port") or 993)
                rua_ssl = _truthy_yes(row.get("imap_ssl"))
                rua_folder = str(row.get("imap_folder") or "INBOX").strip()
            else:
                rua_host = str(row.get("rua_imap_host") or "").strip()
                rua_user = str(row.get("rua_imap_username") or "").strip()
                rua_pass = str(row.get("rua_imap_password") or "").strip()
                rua_port = int(row.get("rua_imap_port") or 993)
                rua_ssl = _truthy_yes(row.get("rua_imap_ssl"))
                rua_folder = str(row.get("rua_imap_folder") or "INBOX").strip()
                if not rua_host or not rua_user or not rua_pass:
                    _fail("same_mailbox_for_rua_and_bounce=no, ale brak danych rua_imap_*")
                    continue

            selectors = _selectors_from_form(str(row.get("dkim_selectors") or ""))
            notes = str(row.get("notes") or "").strip()

            domain_id = domain_repo.create(name=domain_name, dkim_selectors=selectors, notes=notes, enabled=True)

            mailbox_repo.create(
                name=bounce_name, kind="bounce",
                host=host, port=int(row.get("imap_port") or 993), ssl=_truthy_yes(row.get("imap_ssl")),
                username=username, folder=str(row.get("imap_folder") or "INBOX").strip(),
                processed_folder=None, domain=domain_name, enabled=True,
                password_encrypted=encrypt_secret(password),
            )
            existing_mailbox_names.add(bounce_name)

            mailbox_repo.create(
                name=rua_name, kind="rua",
                host=rua_host, port=rua_port, ssl=rua_ssl,
                username=rua_user, folder=rua_folder,
                processed_folder=None, domain=None, enabled=True,
                password_encrypted=encrypt_secret(rua_pass),
            )
            existing_mailbox_names.add(rua_name)

            folder_names = [f.strip() for f in str(row.get("folders") or "").split(",") if f.strip()]
            if folder_names:
                folder_ids = [folder_repo.get_or_create_by_name(fn) for fn in folder_names]
                folder_repo.set_domain_folders(domain_id, folder_ids)

            monitored_names.add(domain_name)
            results.append({"row": row_num, "domain": domain_name, "status": "created", "reason": ""})
        except SecretsError:
            _fail("SECRET_KEY nie jest ustawiony — nie można zaszyfrować hasła")
        except Exception as exc:  # noqa: BLE001 — one bad row must not kill the rest of the import
            _fail(f"nieoczekiwany błąd: {exc}")

    return _render(request, "settings.html", lang, _settings_context(notice="bulk_import_done", bulk_results=results))


@app.post("/settings/domains")
def save_domain(
    request: Request,
    domain_id: Optional[int] = Form(None),
    name: str = Form(...),
    dkim_selectors: str = Form(""),
    notes: str = Form(""),
    enabled: Optional[str] = Form(None),
    # Optional inline "also add a bounce mailbox for this domain" — only
    # offered when adding a brand-new domain (domain_id is None). A domain has
    # at most one natural bounce mailbox, unlike rua, which is shared across
    # every domain and so is never asked for here — see the module docstring.
    add_bounce: Optional[str] = Form(None),
    bounce_name: str = Form(""),
    bounce_host: str = Form(""),
    bounce_port: int = Form(993),
    bounce_ssl: Optional[str] = Form(None),
    bounce_username: str = Form(""),
    bounce_password: str = Form(""),
    bounce_folder: str = Form("INBOX"),
    bounce_processed_folder: str = Form(""),
    # Optional inline "also add the central rua mailbox" — the template only
    # offers this checkbox while zero rua mailboxes exist yet (see
    # settings.html's offer_rua), since rua is shared across every domain and
    # should not be re-prompted for on each subsequent one.
    add_rua: Optional[str] = Form(None),
    rua_name: str = Form(""),
    rua_host: str = Form(""),
    rua_port: int = Form(993),
    rua_ssl: Optional[str] = Form(None),
    rua_username: str = Form(""),
    rua_password: str = Form(""),
    rua_folder: str = Form("INBOX"),
    rua_processed_folder: str = Form(""),
):
    lang = _resolve_request_lang(request)
    ctx = _context()
    domain_repo = DomainConfigRepository(ctx["database"], ctx["settings"].project_id)
    mailbox_repo = MailboxConfigRepository(ctx["database"], ctx["settings"].project_id)

    clean_name = (name or "").strip().lower()
    if not clean_name:
        return _settings_redirect(lang, error="required")

    selectors = _selectors_from_form(dkim_selectors)
    is_enabled = enabled is not None
    is_new = not domain_id

    # --- Validate everything before writing anything ------------------------
    # Two repositories are involved; without a shared transaction, the only
    # way to avoid a half-created domain-with-no-mailbox is to check both
    # halves up front and only then touch the database.
    if is_new and domain_repo.get_by_name(clean_name):
        return _settings_redirect(lang, error="domain_exists")

    want_bounce = is_new and add_bounce is not None
    bounce_fields: Optional[Dict[str, Any]] = None
    if want_bounce:
        if not bounce_host.strip() or not bounce_username.strip():
            return _settings_redirect(lang, error="required")

        clean_bounce_name = bounce_name.strip() or clean_name
        if any(m["name"] == clean_bounce_name for m in mailbox_repo.list_all()):
            return _settings_redirect(lang, error="mailbox_exists")

        bounce_fields = {
            "name": clean_bounce_name,
            "kind": "bounce",
            "host": bounce_host.strip(),
            "port": int(bounce_port),
            "ssl": bounce_ssl is not None,
            "username": bounce_username.strip(),
            "folder": (bounce_folder or "INBOX").strip(),
            "processed_folder": bounce_processed_folder.strip() or None,
            "domain": clean_name,
            "enabled": True,
        }
        if bounce_password:
            try:
                bounce_fields["password_encrypted"] = encrypt_secret(bounce_password)
            except SecretsError:
                return _settings_redirect(lang, error="no_secret_key")

    # Rua has no domain field on the mailbox row — a report's own XML content
    # decides which domain it belongs to, not which mailbox it arrived in
    # (see health.py). This is just a convenience to create one alongside the
    # domain in the same submit; one rua mailbox per domain is as valid a
    # setup as a single consolidated one.
    want_rua = is_new and add_rua is not None
    rua_fields: Optional[Dict[str, Any]] = None
    if want_rua:
        if not rua_host.strip() or not rua_username.strip():
            return _settings_redirect(lang, error="required")

        # A plain default of clean_name would collide with the bounce
        # mailbox's own default of clean_name if both boxes are checked and
        # both names are left blank — the -rua suffix keeps that the common
        # case rather than an error the user has to work around.
        clean_rua_name = rua_name.strip() or f"{clean_name}-rua"
        existing_names = {m["name"] for m in mailbox_repo.list_all()}
        if clean_rua_name in existing_names:
            return _settings_redirect(lang, error="mailbox_exists")
        # Both sub-forms could pick the same default/typed name; catch that
        # collision too, since neither has been written yet at this point.
        if bounce_fields and bounce_fields["name"] == clean_rua_name:
            return _settings_redirect(lang, error="mailbox_exists")

        rua_fields = {
            "name": clean_rua_name,
            "kind": "rua",
            "host": rua_host.strip(),
            "port": int(rua_port),
            "ssl": rua_ssl is not None,
            "username": rua_username.strip(),
            "folder": (rua_folder or "INBOX").strip(),
            "processed_folder": rua_processed_folder.strip() or None,
            "domain": None,
            "enabled": True,
        }
        if rua_password:
            try:
                rua_fields["password_encrypted"] = encrypt_secret(rua_password)
            except SecretsError:
                return _settings_redirect(lang, error="no_secret_key")

    # --- Now write ------------------------------------------------------------
    if domain_id:
        domain_repo.update(
            domain_id, name=clean_name, dkim_selectors=selectors, notes=notes, enabled=is_enabled
        )
    else:
        domain_repo.create(name=clean_name, dkim_selectors=selectors, notes=notes, enabled=is_enabled)
        if bounce_fields:
            mailbox_repo.create(**bounce_fields)
        if rua_fields:
            mailbox_repo.create(**rua_fields)

    return _settings_redirect(lang, notice="domain_saved")


@app.post("/settings/domains/{domain_id}/delete")
def delete_domain(request: Request, domain_id: int):
    lang = _resolve_request_lang(request)
    ctx = _context()
    DomainConfigRepository(ctx["database"], ctx["settings"].project_id).delete(domain_id)
    return _settings_redirect(lang, notice="domain_deleted")


@app.post("/settings/mailboxes")
def save_mailbox(
    request: Request,
    mailbox_id: Optional[int] = Form(None),
    name: str = Form(...),
    kind: str = Form("rua"),
    host: str = Form(...),
    port: int = Form(993),
    ssl: Optional[str] = Form(None),
    username: str = Form(...),
    password: str = Form(""),
    folder: str = Form("INBOX"),
    processed_folder: str = Form(""),
    domain: str = Form(""),
    enabled: Optional[str] = Form(None),
):
    lang = _resolve_request_lang(request)
    ctx = _context()
    repo = MailboxConfigRepository(ctx["database"], ctx["settings"].project_id)

    clean_name = (name or "").strip()
    if not clean_name or not host.strip() or not username.strip():
        return _settings_redirect(lang, error="required")

    clean_domain = domain.strip().lower() or None

    if kind == "bounce":
        if not clean_domain:
            return _settings_redirect(lang, error="required")
        # A bounce mailbox that references a domain not on the monitored list
        # would silently attribute its bounces to that domain, and nothing on
        # the dashboard would show them (there is no domain page for it).
        # Refuse now, when the user is still looking at the form.
        domain_repo = DomainConfigRepository(ctx["database"], ctx["settings"].project_id)
        monitored = {d["name"] for d in domain_repo.list_all()}
        if not monitored:
            return _settings_redirect(lang, error="no_domains_for_bounce")
        if clean_domain not in monitored:
            return _settings_redirect(lang, error="bounce_domain_unknown")

    fields: Dict[str, Any] = {
        "name": clean_name,
        "kind": kind if kind in ("rua", "bounce") else "rua",
        "host": host.strip(),
        "port": int(port),
        "ssl": ssl is not None,
        "username": username.strip(),
        "folder": (folder or "INBOX").strip(),
        "processed_folder": processed_folder.strip() or None,
        "domain": clean_domain,
        "enabled": enabled is not None,
    }

    # An empty password field on an edit means "keep what is stored", so it is
    # only written when the user actually typed something.
    if password:
        try:
            fields["password_encrypted"] = encrypt_secret(password)
            # A freshly typed password supersedes any .env reference.
            fields["password_env"] = None
        except SecretsError:
            return _settings_redirect(lang, error="no_secret_key")

    if mailbox_id:
        repo.update(mailbox_id, **fields)
    else:
        existing = [m for m in repo.list_all() if m["name"] == clean_name]
        if existing:
            return _settings_redirect(lang, error="mailbox_exists")
        repo.create(**fields)

    return _settings_redirect(lang, notice="mailbox_saved")


@app.post("/settings/mailboxes/{mailbox_id}/delete")
def delete_mailbox(request: Request, mailbox_id: int):
    lang = _resolve_request_lang(request)
    ctx = _context()
    MailboxConfigRepository(ctx["database"], ctx["settings"].project_id).delete(mailbox_id)
    return _settings_redirect(lang, notice="mailbox_deleted")


@app.post("/api/test/mailbox")
def api_test_mailbox(
    request: Request,
    mailbox_id: Optional[int] = Form(None),
    # Defaults rather than Form(...): FastAPI treats an empty form field as a
    # missing one and answers 422, which would surface as a blank error in the
    # UI instead of "fill in the required fields".
    host: str = Form(""),
    port: int = Form(993),
    ssl: Optional[str] = Form(None),
    username: str = Form(""),
    password: str = Form(""),
    folder: str = Form("INBOX"),
) -> JSONResponse:
    """Test IMAP credentials from the form, before anything is saved."""
    lang = _resolve_request_lang(request)
    ctx = _context()
    settings, database = ctx["settings"], ctx["database"]
    repo = MailboxConfigRepository(database, settings.project_id)

    if not host.strip() or not username.strip():
        return JSONResponse({"ok": False, "message": translate("test.fields_missing", lang)})

    stored_encrypted = None
    stored_env = None
    if mailbox_id and not password:
        # Editing an existing mailbox without retyping the password: test the
        # one already stored rather than refusing.
        row = repo.get(mailbox_id)
        if row:
            stored_encrypted = row.get("password_encrypted")
            stored_env = row.get("password_env")

    candidate = Mailbox(
        name="test",
        host=host.strip(),
        port=int(port),
        ssl=ssl is not None,
        username=username.strip(),
        folder=(folder or "INBOX").strip(),
        password_encrypted=stored_encrypted,
        password_env=stored_env,
    )

    result = test_mailbox(candidate, password=password or None)
    if mailbox_id:
        repo.record_test(
            mailbox_id,
            ok=result["ok"],
            error=None if result["ok"] else translate(result["message_key"], "en", **result["params"]),
        )

    return JSONResponse(
        {"ok": result["ok"], "message": translate(result["message_key"], lang, **result["params"])}
    )


@app.post("/api/test/domain")
def api_test_domain(
    request: Request, name: str = Form(""), dkim_selectors: str = Form("")
) -> JSONResponse:
    """Check a domain resolves and has the records the tool depends on."""
    lang = _resolve_request_lang(request)
    if not name.strip():
        return JSONResponse({"ok": False, "message": translate("test.fields_missing", lang)})

    result = test_domain(name, _selectors_from_form(dkim_selectors))
    return JSONResponse(
        {
            "ok": result["ok"],
            "warning": result.get("warning", False),
            "message": translate(result["message_key"], lang, **result["params"]),
        }
    )


# --- Run now ------------------------------------------------------------------


@app.post("/api/run/{stream}")
def api_run_stream(request: Request, stream: str) -> JSONResponse:
    """Trigger one ingestion stream on demand."""
    lang = _resolve_request_lang(request)
    ctx = _context()
    result = jobs.start(stream, ctx["settings"], ctx["database"])

    if result["started"]:
        message = translate("action.started", lang, stream=translate(f"stream.{stream}", lang))
    elif result["reason"] == "already_running":
        message = translate("action.already_running", lang, stream=translate(f"stream.{stream}", lang))
    else:
        message = result["reason"] or ""

    return JSONResponse({**result, "message": message})


@app.get("/api/run/status")
def api_run_status() -> JSONResponse:
    """Which streams are mid-run, for the UI to poll after pressing the button."""
    ctx = _context()
    return JSONResponse({"running": jobs.running_streams(ctx["settings"], ctx["database"])})


@app.get("/api/health")
def api_health(lang: str = DEFAULT_LANG) -> JSONResponse:
    """Ingestion freshness, for scripting or an external monitor.

    Defaults to Polish (the dashboard's default) but any consumer can ask for
    English with ?lang=en.
    """
    ctx = _context()
    rows = localize_ingestion_health(ingestion_health(ctx["database"], ctx["settings"]), resolve_lang(lang))
    return JSONResponse(
        [
            {
                **{k: v for k, v in row.items() if k not in ("message_params",)},
                "last_success": row["last_success"].isoformat() if row["last_success"] else None,
            }
            for row in rows
        ]
    )
